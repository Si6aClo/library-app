from pathlib import Path

from aiosqlite import Connection
from fastapi import APIRouter, Depends
from openpyxl import Workbook
from redmail import outlook
from starlette import status
from starlette.background import BackgroundTasks

from config import get_settings
from db import get_cursor

api_router = APIRouter(tags=["Email sender"])

async def send_mail(cursor: Connection = Depends(get_cursor), ):
    res = await cursor.execute("SELECT * FROM category")
    res = await res.fetchall()
    all_categories = {item[0]: item[1] for item in res}

    res = await cursor.execute("SELECT * FROM book")
    res = await res.fetchall()

    wb = Workbook()
    all_using_letters = ['A', 'B', 'C', 'D', 'E', 'F']
    ws = wb.active
    ws['A1'] = "Номер книги"
    ws['B1'] = "Категория"
    ws['C1'] = "Название книги"
    ws['D1'] = "Описание книги"
    ws['E1'] = "Автор"
    ws['F1'] = "Количество"

    row = 1
    for item in res:
        row += 1
        for i in range(len(all_using_letters)):
            ws[f'{all_using_letters[i]}{row}'] = item[i] if i != 1 else all_categories[item[i]]

    wb.save("library.xlsx")
    wb.close()

    outlook.username = get_settings().OUTLOOK_USER
    outlook.password = get_settings().OUTLOOK_PASSWORD

    outlook.send(
        subject="Данные Библиотеки",
        receivers=[get_settings().SEND_TO],
        text="Приветствую, вот ваша excel таблица с информацией о всех книгах!",
        attachments={
            "library.xlsx": Path("./library.xlsx"),
        }
    )


@api_router.post(
    "/email_send/send",
    status_code=status.HTTP_200_OK,
)
async def email_sender(background_tasks: BackgroundTasks, cursor: Connection = Depends(get_cursor), ):
    background_tasks.add_task(send_mail, cursor)
